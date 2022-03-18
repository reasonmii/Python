
# 엑셀 파일 생성해서 내용 작성하기

import os
import random
import openpyxl

# Create an excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "info"                          # Sheet name
ws.cell(row=1, column=1).value = "hello"   # Cell contents
wb.save("test.xlsx")                       # File name

ws.cell(row=1, column=1).value = "name"
ws.cell(row=1, column=2).value = "phone number"

# Contents
last_names = ["Kim", "Lee", "Park", "Shin", "Song", "Hwang", "Jang", "Choi"]
names = ["Jully", "Michelle", "Mike", "Cathy", "Lisa", "John", "James", "Tylor", "Ari", "Emma"]
phone_numbers = ["010-1234-5678", "010-8765-4321"]

# customer_list_1 ~ customer_list_49 까지 excel 파일 생성
# 1열 : 이름, 2열 : 전화번호
for excel_idx in range(50):
    for row in range(2, 50):   # 첫 번째 줄은 name, phone number 라는 header 들어감
        # random.choice : get any value randomly from the list
        ws.cell(row=row, column=1).value = random.choice(last_names) + random.choice(names)
        ws.cell(row=row, column=2).value = random.choice(phone_numbers)
        
    wb.save(f"./customer_list_{excel_idx}.xlsx")
